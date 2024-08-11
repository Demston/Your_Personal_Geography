"""Скрипт создания карты с отметками на основе таблицы Excel"""

from create_files import *      # скрипт с созданием отсутствующих файлов
import os
import folium                              # работа с картой
from folium.plugins import MarkerCluster   # маркеры на карте
from geopy.geocoders import Nominatim      # работа с координатами
import openpyxl as xl                      # excel
from tkinter import filedialog
import tkinter.messagebox as mb

create_map()
create_table()
create_map_path()
create_table_path()

table_path = ''
map_path = ''
warning_list = ''


def open_file():
    """Окно открытия файла с таблицей"""
    global table_path
    table_path = filedialog.askopenfilename()
    open('table_path.txt', 'w').write(table_path)
    os.startfile(table_path)
    return table_path


def open_current_file():
    """Окно открытия текущего файла с таблицей"""
    global table_path
    table_path = open('table_path.txt').read()
    os.startfile(table_path)


def select_file():
    """Окно выбора актуальной таблицы"""
    global table_path
    table_path = filedialog.askopenfilename()
    open('table_path.txt', 'w').write(table_path)
    return table_path


def show_warning():
    """Предупреждение"""
    global warning_list
    if len(warning_list) != 0:
        msg = f"        Wrong adresses:        \n\n{warning_list}"
        mb.showwarning("Warning", msg)
    else:
        pass


def save_map():
    """Сохранить карту"""
    global map_path
    if len(open('map_path.txt').read()) == 0:
        map_path = os.getcwd() + '\\map.html'
        open('map_path.txt', 'w').write(map_path)
        my_map.save(map_path)
    else:
        open(map_path, 'w')
        my_map.save("map.html")


def save_map_as():
    """Сохранить карту как"""
    global map_path
    path_save_map = filedialog.asksaveasfilename(title="Сохранить файл", defaultextension=".html",
                                                 filetypes=[("Веб-страница", "*.html")])
    map_path = open('map_path.txt', 'w').write(path_save_map)
    my_map.save(path_save_map)


table_file = open('table_path.txt', 'r').read()   # читаем файл с путём к таблице
if len(table_file) != 0:  # если в файле прописан путь
    if os.path.exists(table_file) is True:  # если такой есть
        table_path = table_file
        places_file = xl.load_workbook(table_path)  # файл
        read_column_adr = places_file['places_page']  # страница
    else:   # если такого нет
        open_file()
        # table_path = os.getcwd()+'\\places.xlsx'
        table_file = open('table_path.txt', 'w').write(table_path)
        places_file = xl.load_workbook(table_path)  # файл
        read_column_adr = places_file['places_page']  # страница
else:  # если путь не прописан
    default_file_path = os.getcwd()+'\\places.xlsx'
    table_file = open('table_path.txt', 'w').write(default_file_path)
    places_file = xl.load_workbook(default_file_path)  # файл
    read_column_adr = places_file['places_page']  # страница

places_id = []  # id по номеру строки excel
places_list = []  # список адресов
desc_places = []  # названия объектов
object_color = []   # цвет объекта на карте

count_id = 0
for a in read_column_adr['B'][3:]:
    places_id.append(count_id)  # добавим id в список
    count_id += 1
for b in read_column_adr['B'][3:]:
    places_list.append(b.value)  # добавим адреса мест в список
for c in read_column_adr['C'][3:]:
    desc_places.append(c.value)  # добавим названия объектов в список
for d in read_column_adr['D'][3:]:
    if d.value is None:
        object_color.append('white')  # добавим отображаемый цвет объекта в список
    else:
        object_color.append(d.value)  # добавим отображаемый цвет объекта в список

places_dict = dict(zip(places_id, places_list))  # сформиуем словарь всех адресов и координат
desc_places = dict(zip(places_id, desc_places))  # сформиуем словарь описания объектов
colors_dict = dict(zip(places_id, object_color))   # сформируем словарь раскраски объектов

adr_from_file_dict = {}  # словарь адресов из файла
crd_from_file_dict = {}  # словарь координат из файла
crd_from_adr_dict = {}  # словарь координат на основе адресов

# рассортируем места по словарям
for key, value in places_dict.items():
    try:
        for el in value:
            if el[0].isdigit():  # по признаку цифр (координаты)
                crd_from_file_dict[key] = value
                break
            else:  # по признаку букв (буквы, адрес)
                adr_from_file_dict[key] = value
    except TypeError:
        continue

app = Nominatim(user_agent="tutorial")

# узнаём координаты по адресу и добавляем в словарь
for key, value in adr_from_file_dict.items():
    try:
        location = app.geocode(value).raw
        crds = (location['lat'] + ', ' + location['lon'])
        crd_from_adr_dict[key] = crds
    except AttributeError:
        warning_list += ('-  '+value+'\n')
        # print(f'Поправь адрес места: {value}')

crd_all_str_dict = crd_from_file_dict | crd_from_adr_dict  # словарь всех координат (текст)
crd_all_float_dict = {}  # словарь всех координат (цифры)

# делим строку на широту и долготу, переводим в цифры, добавляем в словарь
for key, value in crd_all_str_dict.items():
    try:
        a = float(value.split(', ')[0])
        b = float(value.split(', ')[-1])
        c = [a, b]
        crd_all_float_dict[key] = c
    except ValueError:
        warning_list += '•  '+value+'\n'
        # print(f'Поправь адрес места: {value}')

try:
    my_map = folium.Map(location=crd_all_float_dict[0], zoom_start=12)  # начальная точка отображения
except KeyError:
    my_map = folium.Map(location=[55, 55], zoom_start=5)  # если с первой точкой что-то не так

marker_cluster = MarkerCluster().add_to(my_map)  # опционально, группировка точек

for ids, coordinates in crd_all_float_dict.items():  # рисуем точки на карте
    folium.CircleMarker(location=coordinates, popup=str(desc_places[ids]), radius=7, fill_color=str(colors_dict[ids]),
                        color='gray', fill_opacity=0.8).add_to(my_map)  # или add_to(marker_cluster)

if __name__ == '__main__':
    show_warning()
    my_map.save("map.html")
    # print('Карта готова')
